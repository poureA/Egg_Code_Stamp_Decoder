import openpyxl as op
class Ecsd(object):
    '''class docstring'''
    def __init__(self,stamp_code)->None:
        self._sc = stamp_code
    def Decode(self)->None:
        '''function docstring'''
        farming_method = self._sc[0]
        country_of_origin = self._sc[1:3]
        farm_id = self._sc[3:]
        if farming_method == '0':
            farming_method='Organic'
        elif farming_method == '1':
            farming_method='Free range'
        elif farming_method == '2':
            farming_method='Barn'
        elif farming_method == '3':
            farming_method='Cage'
        else :
            return f'farming method must be 0 , 1 , 2 or 3 not {farming_method}'
        work_book = op.load_workbook('D:\\text\\countries.xlsx')
        sheet = work_book.active
        c_names = []
        Abbreviation = []
        for i in range(2,sheet.max_row):
            country = sheet.cell(row=i,column=1).value
            c_names.append(country)
            abb = sheet.cell(row=i,column=2).value
            Abbreviation.append(abb)
        if country_of_origin in Abbreviation :
            idx = Abbreviation.index(country_of_origin)
            country_of_origin=c_names[idx]
        else :
            return f'{country_of_origin} is not a country !'
        if farm_id.isdigit() is False :
            return '%s is not a true farm id'%(farm_id)
        return '{0} egg\nCountry of Origin: {1}\nFarm Id: {2}'.format(farming_method,country_of_origin,farm_id)
ask = input('enter a code stamp :')
if len(ask)>=7:
    egg = Ecsd(ask)
    print(egg.Decode())
else :
    print('A valid code contains at least 7 alphanumerical characters')
exit = input('please enter any key to exit :')
